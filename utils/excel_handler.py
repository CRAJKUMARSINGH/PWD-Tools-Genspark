"""
Excel Handler Utility for PWD Tools Desktop Application
Handles Excel file operations and data processing
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os
from pathlib import Path

class ExcelHandler:
    """Utility class for Excel file operations"""
    
    def __init__(self):
        self.supported_formats = ['.xlsx', '.xls']
    
    def read_excel(self, file_path, sheet_name=None):
        """Read Excel file and return DataFrame"""
        try:
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(file_path)
            return df
        except Exception as e:
            print(f"Error reading Excel file: {e}")
            return None
    
    def write_excel(self, data, file_path, sheet_name='Sheet1'):
        """Write data to Excel file"""
        try:
            if isinstance(data, pd.DataFrame):
                data.to_excel(file_path, sheet_name=sheet_name, index=False)
            else:
                # Convert list/dict to DataFrame
                df = pd.DataFrame(data)
                df.to_excel(file_path, sheet_name=sheet_name, index=False)
            return True
        except Exception as e:
            print(f"Error writing Excel file: {e}")
            return False
    
    def create_formatted_excel(self, data, file_path, sheet_name='Sheet1', title=None):
        """Create formatted Excel file with styling"""
        try:
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # Add title if provided
            if title:
                ws['A1'] = title
                ws['A1'].font = Font(size=16, bold=True)
                ws.merge_cells('A1:D1')
                ws['A1'].alignment = Alignment(horizontal='center')
                ws.row_dimensions[1].height = 30
                
                # Start data from row 3
                start_row = 3
            else:
                start_row = 1
            
            # Convert data to rows
            if isinstance(data, pd.DataFrame):
                rows = dataframe_to_rows(data, index=False, header=True)
            else:
                df = pd.DataFrame(data)
                rows = dataframe_to_rows(df, index=False, header=True)
            
            # Add data to worksheet
            for r_idx, row in enumerate(rows, start_row):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    
                    # Style header row
                    if r_idx == start_row:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                        cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=start_row):
                for cell in row:
                    cell.border = thin_border
            
            # Save workbook
            wb.save(file_path)
            return True
            
        except Exception as e:
            print(f"Error creating formatted Excel: {e}")
            return False
    
    def validate_excel_structure(self, file_path, required_columns):
        """Validate Excel file has required columns"""
        try:
            df = self.read_excel(file_path)
            if df is None:
                return False, "Could not read Excel file"
            
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return False, f"Missing required columns: {', '.join(missing_columns)}"
            
            return True, "File structure is valid"
            
        except Exception as e:
            return False, f"Error validating file: {str(e)}"
    
    def get_sheet_names(self, file_path):
        """Get list of sheet names from Excel file"""
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            return wb.sheetnames
        except Exception as e:
            print(f"Error getting sheet names: {e}")
            return []
    
    def create_summary_sheet(self, data, file_path, summary_title="Summary"):
        """Create Excel file with summary sheet"""
        try:
            wb = openpyxl.Workbook()
            
            # Summary sheet
            summary_ws = wb.active
            summary_ws.title = "Summary"
            
            # Add summary information
            summary_ws['A1'] = summary_title
            summary_ws['A1'].font = Font(size=16, bold=True)
            
            summary_ws['A3'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            summary_ws['A4'] = f"Total Records: {len(data) if isinstance(data, (list, pd.DataFrame)) else 'N/A'}"
            
            # Data sheet
            if isinstance(data, pd.DataFrame):
                data_ws = wb.create_sheet("Data")
                for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        data_ws.cell(row=r_idx, column=c_idx, value=value)
            
            wb.save(file_path)
            return True
            
        except Exception as e:
            print(f"Error creating summary sheet: {e}")
            return False
    
    def export_to_multiple_sheets(self, data_dict, file_path):
        """Export multiple datasets to different sheets in one Excel file"""
        try:
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            for sheet_name, data in data_dict.items():
                ws = wb.create_sheet(sheet_name)
                
                if isinstance(data, pd.DataFrame):
                    for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
                        for c_idx, value in enumerate(row, 1):
                            ws.cell(row=r_idx, column=c_idx, value=value)
                else:
                    # Handle list/dict data
                    df = pd.DataFrame(data)
                    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                        for c_idx, value in enumerate(row, 1):
                            ws.cell(row=r_idx, column=c_idx, value=value)
            
            wb.save(file_path)
            return True
            
        except Exception as e:
            print(f"Error exporting to multiple sheets: {e}")
            return False